# ---------------- Full CNN-EQIC + INN Streamlit App (Manual Entry Inference Mode) ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import joblib
import io
import base64
from sklearn.linear_model import LinearRegression

# ---------------- Utility Functions ----------------
def interpret_level(value):
    if value < 0.33:
        return "low"
    elif value < 0.66:
        return "moderate"
    else:
        return "high"

def generate_cluster_descriptions(centroids, feature_names):
    descriptions = []
    for i, centroid in enumerate(centroids):
        traits = []
        for feat_name, val in zip(feature_names, centroid):
            level = interpret_level(val)
            traits.append(f"{level} {feat_name}")
        descriptions.append(f"Cluster {i} is characterized by {', '.join(traits)}.")
    return descriptions

def generate_comparative_summary(centroids, feature_names):
    ranges = centroids.max(axis=0) - centroids.min(axis=0)
    important_idx = np.where(ranges > 0.2)[0]
    if len(important_idx) == 0:
        return "Clusters show similar profiles."
    lines = ["Comparative summary of cluster differences:"]
    for idx in important_idx:
        feat = feature_names[idx]
        vals = centroids[:, idx]
        highs = np.where(vals > 0.66)[0]
        lows = np.where(vals < 0.33)[0]
        lines.append(f"- '{feat}' is high in clusters {list(highs)} and low in clusters {list(lows)}")
    return "\n".join(lines)

def load_scaler_from_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(row[0]) for row in scaler_sheet.iter_rows(values_only=True) if row[0]]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    return joblib.load(io.BytesIO(scaler_bytes))

def manually_create_lagged_input(input_dict, window_size):
    data = {}
    for feat, val in input_dict.items():
        for t in range(window_size):
            data[f"{feat}_t{t}"] = val
    return pd.DataFrame([data])

# ---------------- Streamlit App ----------------
st.set_page_config(page_title="CNN-EQIC + INN", layout="wide")
st.title("\U0001F4CA CNN-EQIC Clustering & \U0001F9E0 INN Inference")
tabs = st.tabs(["CNN-EQIC Clustering", "INN Inference Calculator"])

with tabs[0]:
    st.header("CNN-EQIC Clustering")
    file = st.file_uploader("Upload training dataset", type=["csv", "xlsx"])
    if file:
        df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
        numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        selected = st.multiselect("Select features", numerical_cols, default=numerical_cols[:3])
        window = st.slider("Lag window", 2, 10, 3)
        if len(selected) >= 2:
            clean = SimpleImputer().fit_transform(df[selected])
            scaler = MinMaxScaler().fit(clean)
            scaled = scaler.transform(clean)
            patterns = [scaled[i - window:i].flatten() for i in range(window, len(scaled))]
            feature_names = [f"{f}_t{t}" for t in range(window) for f in selected]
            k = st.slider("Number of clusters", 2, 10, 4)
            km = KMeans(n_clusters=k, random_state=0).fit(patterns)
            centroids = km.cluster_centers_

            st.subheader("Cluster Descriptions")
            for desc in generate_cluster_descriptions(centroids, feature_names):
                st.markdown(f"- {desc}")

            st.subheader("Comparative Summary")
            st.text(generate_comparative_summary(centroids, feature_names))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Centroids"
            df_c = pd.DataFrame(centroids, columns=feature_names)
            for r in dataframe_to_rows(df_c, index=False, header=True):
                ws.append(r)
            scaler_io = io.BytesIO()
            joblib.dump(scaler, scaler_io)
            b64 = base64.b64encode(scaler_io.getvalue()).decode("utf-8")
            sc_sheet = wb.create_sheet("Scaler")
            for i in range(0, len(b64), 1000):
                sc_sheet.append([b64[i:i+1000]])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button("Download CNN-EQIC Output", output, "CNN_EQIC_output.xlsx")

with tabs[1]:
    st.header("INN Inference Calculator")
    uploaded = st.file_uploader("Upload CNN-EQIC Output (Excel)", type="xlsx")
    window = st.number_input("Lag window size used in CNN-EQIC", 2, 10, 3)

    if uploaded:
        df_centroids = pd.read_excel(uploaded, sheet_name="Centroids")
        scaler = load_scaler_from_excel(uploaded)
        base_features = sorted(set(col.rsplit("_t", 1)[0] for col in df_centroids.columns))
        target = st.selectbox("Select variable to infer:", base_features)
        required_inputs = [f for f in base_features if f != target]

        st.subheader("Input values for required features")
        inputs = {}
        for feat in required_inputs:
            inputs[feat] = st.number_input(f"{feat}", value=0.0)

        if st.button("Generate Inference"):
            lagged_df = manually_create_lagged_input(inputs, window)
            used_cols = [col for col in df_centroids.columns if not col.startswith(target + "_t")]
            target_cols = [col for col in df_centroids.columns if col.startswith(target + "_t")]

            result = {}
            for tgt_col in target_cols:
                model = LinearRegression()
                model.fit(df_centroids[used_cols], df_centroids[tgt_col])
                pred = model.predict(lagged_df[used_cols])[0]
                result[tgt_col] = pred

            # Prepare inverse scaled output
            values = np.array([[result[col] for col in target_cols]])
            dummy_df = pd.DataFrame(np.zeros((1, len(df_centroids.columns))), columns=df_centroids.columns)
            for col in target_cols:
                dummy_df[col] = result[col]
            unscaled = scaler.inverse_transform(dummy_df)[0]
            st.subheader("Inference Result")
            for col in target_cols:
                st.write(f"{col} (scaled): {result[col]:.4f}  |  unscaled: {unscaled[df_centroids.columns.get_loc(col)]:.4f}")
