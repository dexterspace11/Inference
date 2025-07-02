# ---------------- Full CNN-EQIC + INN Streamlit App ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from scipy.cluster.hierarchy import linkage, dendrogram
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import joblib
import io
import base64
from sklearn.linear_model import LinearRegression

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Narrative Generator ----------------
def interpret_level(value):
    if value < 0.33:
        return "low"
    elif value < 0.66:
        return "moderate"
    else:
        return "high"

def generate_cluster_descriptions(centroids, feature_names):
    descriptions = []
    for i, centroid in enumerate(centroids):
        traits = []
        for feat_name, val in zip(feature_names, centroid):
            level = interpret_level(val)
            traits.append(f"{level} {feat_name}")
        trait_str = ", ".join(traits)
        desc = f"Cluster {i} is characterized by {trait_str}."
        descriptions.append(desc)
    return descriptions

def generate_comparative_summary(centroids, feature_names):
    feature_ranges = centroids.max(axis=0) - centroids.min(axis=0)
    important_features_idx = np.where(feature_ranges > 0.2)[0]
    if len(important_features_idx) == 0:
        return "Clusters show similar profiles with minimal variation."

    lines = ["Comparative summary of cluster differences:"]
    for idx in important_features_idx:
        feat = feature_names[idx] if idx < len(feature_names) else f"Feature_{idx}"
        vals = centroids[:, idx]
        high_clusters = np.where(vals > 0.66)[0]
        low_clusters = np.where(vals < 0.33)[0]
        line = f"- '{feat}' is high in clusters {list(high_clusters)} and low in clusters {list(low_clusters)}"
        lines.append(line)
    return "\n".join(lines)

# ---------------- Helper: Load Scaler from 'Scaler' sheet ----------------
def load_scaler_from_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    if "Scaler" not in wb.sheetnames:
        raise ValueError("Scaler sheet not found in Excel file.")
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(cell[0].value) for cell in scaler_sheet.iter_rows(values_only=True) if cell[0].value]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    scaler = joblib.load(io.BytesIO(scaler_bytes))
    return scaler

# ---------------- Helper: Create lagged features from raw test data ----------------
def create_lagged_features(df, features, window_size):
    lagged_data = {}
    for feature in features:
        for lag in range(window_size):
            col_name = f"{feature}_t{lag}"
            lagged_data[col_name] = df[feature].shift(window_size - 1 - lag)
    lagged_df = pd.DataFrame(lagged_data)
    lagged_df.dropna(inplace=True)
    lagged_df.reset_index(drop=True, inplace=True)
    return lagged_df

# ---------------- Streamlit App ----------------
st.set_page_config(page_title="CNN-EQIC + INN Integrated App", layout="wide")
st.title("📊 CNN-EQIC Clustering & 🧠 INN Missing Variable Inference")

tab = st.tabs(["CNN-EQIC Clustering", "INN Inference"])

with tab[0]:
    st.header("CNN-EQIC: Advanced Clustering with Scaler Export")

    uploaded_file = st.file_uploader("Upload raw training data (CSV or Excel) for clustering", type=["csv", "xlsx"], key="cnn_upload")
    if uploaded_file:
        df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
        st.dataframe(df.head())

        numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        default_selection = numerical_cols[:min(4, len(numerical_cols))]
        selected = st.multiselect("Select features", numerical_cols, default=default_selection, key="cnn_features")
        window_size = st.slider("Window size", 2, 20, 5, key="cnn_window")

        if len(selected) >= 2:
            clean = SimpleImputer().fit_transform(df[selected])
            scaler = MinMaxScaler().fit(clean)
            scaled = scaler.transform(clean)

            net = HybridNeuralNetwork(working_memory_capacity=20, decay_rate=100.0)
            for i in range(window_size, len(scaled)):
                pattern = scaled[i - window_size:i].flatten()
                net.process_input(pattern)

            patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
            if len(patterns) > 0:
                patterns = np.array(patterns)
                expanded_feature_names = [f"{feat}_t{t}" for t in range(window_size) for feat in selected]

                k = st.slider("Number of clusters", 2, min(10, len(patterns)), 5, key="cnn_clusters")
                kmeans = KMeans(n_clusters=k, random_state=42).fit(patterns)
                labels = kmeans.labels_
                centroids = kmeans.cluster_centers_

                silhouette = silhouette_score(patterns, labels)
                db = davies_bouldin_score(patterns, labels)
                ch = calinski_harabasz_score(patterns, labels)
                st.metric("Silhouette Score", f"{silhouette:.3f}")
                st.metric("Davies-Bouldin Index", f"{db:.3f}")
                st.metric("Calinski-Harabasz Score", f"{ch:.1f}")

                st.subheader("📌 PCA Cluster Projection")
                pca = PCA(n_components=2).fit_transform(patterns)
                pca_df = pd.DataFrame(pca, columns=['PC1', 'PC2'])
                pca_df['Cluster'] = labels
                fig, ax = plt.subplots()
                sns.scatterplot(data=pca_df, x='PC1', y='PC2', hue='Cluster', palette='tab10', ax=ax)
                st.pyplot(fig)

                st.subheader("🧠 Cluster Interpretations")
                cluster_descriptions = generate_cluster_descriptions(centroids, expanded_feature_names)
                for desc in cluster_descriptions:
                    st.markdown(f"- {desc}")

                st.subheader("🧮 Comparative Feature Summary")
                comp_summary = generate_comparative_summary(centroids, expanded_feature_names)
                st.text(comp_summary)

                st.subheader("📊 Correlation Matrix")
                fig2, ax2 = plt.subplots()
                sns.heatmap(pd.DataFrame(clean, columns=selected).corr(), annot=True, cmap="coolwarm", ax=ax2)
                st.pyplot(fig2)

                st.subheader("🧬 Dendrogram (Hierarchical Clustering)")
                linkage_matrix = linkage(patterns, method='ward')
                fig3, ax3 = plt.subplots(figsize=(10, 4))
                dendrogram(linkage_matrix, truncate_mode='level', p=5, ax=ax3)
                st.pyplot(fig3)

                st.subheader("📘 Academic Report Summary")
                st.markdown("""
                    This clustering analysis segmented the data into meaningful groups based on feature patterns over time.
                    Each cluster reveals typical configurations, such as consistently high or low values in specific features
                    across time windows. Clusters with strong separation (high silhouette scores) indicate clear behavioral
                    differentiation. The dendrogram and PCA plots help visualize these relationships.
                """)

                if st.button("Export CNN-EQIC Results and Scaler to Excel"):
                    wb = openpyxl.Workbook()
                    ws1 = wb.active
                    ws1.title = "Cluster Descriptions"
                    for line in cluster_descriptions:
                        ws1.append([line])
                    ws2 = wb.create_sheet("Comparative Summary")
                    for line in comp_summary.split("\n"):
                        ws2.append([line])
                    ws3 = wb.create_sheet("Cluster Labels")
                    df_assign = pd.DataFrame({"Pattern Index": range(len(labels)), "Cluster": labels})
                    for r in dataframe_to_rows(df_assign, index=False, header=True):
                        ws3.append(r)
                    ws4 = wb.create_sheet("Centroids")
                    df_centroids = pd.DataFrame(centroids, columns=expanded_feature_names)
                    for r in dataframe_to_rows(df_centroids, index=False, header=True):
                        ws4.append(r)
                    # Add raw data by cluster
                    for i in range(k):
                        cluster_data = patterns[labels == i]
                        df_cluster = pd.DataFrame(cluster_data, columns=expanded_feature_names)
                        ws = wb.create_sheet(f"Raw Cluster {i}")
                        for r in dataframe_to_rows(df_cluster, index=False, header=True):
                            ws.append(r)
                    # Save scaler as base64 string in a new sheet "Scaler"
                    scaler_bytes_io = io.BytesIO()
                    joblib.dump(scaler, scaler_bytes_io)
                    scaler_bytes = scaler_bytes_io.getvalue()
                    scaler_b64 = base64.b64encode(scaler_bytes).decode('utf-8')
                    ws_scaler = wb.create_sheet("Scaler")
                    chunk_size = 1000
                    for i in range(0, len(scaler_b64), chunk_size):
                        ws_scaler.append([scaler_b64[i:i+chunk_size]])
                    # Auto-adjust column widths
                    for ws_iter in wb.worksheets:
                        for col_cells in ws_iter.columns:
                            length = max(len(str(cell.value) or "") for cell in col_cells)
                            ws_iter.column_dimensions[col_cells[0].column_letter].width = length + 2
                    # Instead of saving to disk, save to bytes buffer and provide download button:
                    excel_io = io.BytesIO()
                    wb.save(excel_io)
                    excel_io.seek(0)
                    st.download_button(
                        label="Download CNN-EQIC Excel Output",
                        data=excel_io,
                        file_name="CNN_EQIC_output.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

with tab[1]:
    st.header("INN Inference: Missing Variable Imputation")

    cnn_eqic_file = st.file_uploader("Upload CNN-EQIC Excel Output (with scaler)", type=["xlsx"], key="inn_cnn_upload")
    raw_test_file = st.file_uploader("Upload Raw Test Dataset (CSV or Excel)", type=["csv", "xlsx"], key="inn_test_upload")
    output_path = st.text_input("Enter full path to save inference output Excel (e.g., C:/path/to/INNoutput.xlsx)", key="inn_output_path")
    window_size = st.number_input("Enter CNN-EQIC window size used during training", min_value=2, max_value=20, value=5, key="inn_window")

    if cnn_eqic_file and raw_test_file and output_path:
        try:
            # Load centroids and scaler from CNN-EQIC file
            df_centroids = pd.read_excel(cnn_eqic_file, sheet_name="Centroids")
            scaler = load_scaler_from_excel(cnn_eqic_file)

            # Load raw test data
            df_test = pd.read_csv(raw_test_file) if raw_test_file.name.endswith(".csv") else pd.read_excel(raw_test_file)

            # Get base features from centroid columns (strip lag suffixes)
            base_features = sorted(set(col.rsplit('_t', 1)[0] for col in df_centroids.columns))

            # Create lagged features from raw test data to match centroid columns
            df_test_lagged = create_lagged_features(df_test, base_features, window_size)

            # Confirm columns match centroid columns exactly
            missing_cols = set(df_centroids.columns) - set(df_test_lagged.columns)
            if missing_cols:
                st.error(f"Column mismatch: Test data after lagging is missing columns: {missing_cols}")
                st.stop()

            # Identify missing variable base name by checking columns missing in raw test data vs centroid columns base features
            missing_vars = set(base_features) - set(df_test.columns)
            if len(missing_vars) != 1:
                st.error(f"Expected exactly 1 missing variable, but found: {missing_vars}")
                st.stop()
            missing_var = list(missing_vars)[0]
            st.success(f"Missing variable detected for inference: {missing_var}")

            # Prepare features for model training and prediction
            used_columns = [col for col in df_centroids.columns if not col.startswith(missing_var + '_t')]
            target_columns = [col for col in df_centroids.columns if col.startswith(missing_var + '_t')]

            # Check if all used_columns are present in lagged test data
            missing_inputs = set(used_columns) - set(df_test_lagged.columns)
            if missing_inputs:
                st.error(f"Required input columns missing from lagged test data: {missing_inputs}")
                st.stop()

            # Train Linear Regression models for each lag of missing variable
            inferred_data = df_test_lagged.copy()
            for target_col in target_columns:
                model = LinearRegression()
                model.fit(df_centroids[used_columns], df_centroids[target_col])
                inferred_data[target_col] = model.predict(df_test_lagged[used_columns])

            st.subheader("✅ Inferred Values (Scaled)")
            st.write(inferred_data[target_columns].head())

            # Inverse transform inferred values with scaler
            try:
                inferred_scaled = inferred_data[target_columns].values
                inferred_unscaled = scaler.inverse_transform(inferred_scaled)
                for i, col in enumerate(target_columns):
                    inferred_data[col + "_unscaled"] = inferred_unscaled[:, i]
                st.subheader("🎯 Inferred Values (Inverse Transformed)")
                st.write(inferred_data[[col + "_unscaled" for col in target_columns]].head())
            except Exception as scaler_err:
                st.warning(f"⚠️ Inverse transform skipped: {scaler_err}")

            # Save to Excel: raw test data + inferred missing variable (both scaled and unscaled)
            wb = openpyxl.Workbook()
            ws_raw = wb.active
            ws_raw.title = "Raw Test Data"
            for r in dataframe_to_rows(df_test, index=False, header=True):
                ws_raw.append(r)
            ws_infer = wb.create_sheet("INN_Inference")
            for r in dataframe_to_rows(inferred_data, index=False, header=True):
                ws_infer.append(r)

            wb.save(output_path)
            st.success(f"📁 Inference results saved to: {output_path}")

        except Exception as e:
            st.error(f"❌ An error occurred: {e}")
