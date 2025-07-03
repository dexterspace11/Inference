# ---------------- INN Cluster-Based Inference App ----------------
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import base64
import io
from sklearn.metrics.pairwise import euclidean_distances
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="INN - Cluster-Based Inference", layout="wide")
st.title("🧠 INN - Inference Engine for Missing Variable using CNN-EQIC Clusters")

# ---------------- File Upload ----------------
st.header("Step 1: Upload CNN-EQIC Excel Model Output")
model_file = st.file_uploader("Upload CNN-EQIC Excel file", type=["xlsx"])

st.header("Step 2: Upload Test Dataset (with one missing variable)")
test_file = st.file_uploader("Upload test Excel file (with base variable columns)", type=["xlsx"])

st.header("Step 3: Enter Full Output Path")
output_path = st.text_input("Example: C:/Users/yourname/Documents/INNoutput.xlsx")

if model_file and test_file and output_path:
    try:
        # ----- Load model Excel -----
        xl = pd.ExcelFile(model_file)

        # Load centroids and expanded feature names
        df_centroids = pd.read_excel(xl, sheet_name="Centroids")
        expanded_feature_names = [r[0] for r in pd.read_excel(xl, sheet_name="ScalerFeatures", header=None).values.tolist()]

        # Infer base variables and window size
        base_vars = list(set([name.split("_t")[0] for name in expanded_feature_names]))
        window_size = max([int(name.split("_t")[1]) for name in expanded_feature_names]) + 1

        # Load scaler from base64 string
        scaler_b64 = "".join([r[0] for r in pd.read_excel(xl, sheet_name="Scaler", header=None).values.tolist()])
        scaler_bytes = base64.b64decode(scaler_b64.encode("utf-8"))
        scaler = joblib.load(io.BytesIO(scaler_bytes))

        # ----- Load test dataset -----
        df_test_raw = pd.read_excel(test_file)
        test_vars = df_test_raw.columns.tolist()

        # Infer missing variable
        inferred_missing = list(set(base_vars) - set(test_vars))
        if len(inferred_missing) != 1:
            missing_var = st.selectbox("Select missing variable", base_vars)
        else:
            missing_var = inferred_missing[0]
            st.success(f"🔍 Missing variable detected: **{missing_var}**")

        # Prepare test data
        used_vars = [var for var in base_vars if var != missing_var]
        missing_target_cols = [col for col in expanded_feature_names if col.startswith(missing_var + "_")]
        used_expanded_cols = [col for col in expanded_feature_names if not col.startswith(missing_var + "_")]

        # Check test data validity
        if not all(var in df_test_raw.columns for var in used_vars):
            st.error("❌ Your test data is missing required columns for inference.")
            st.stop()

        # ----- Generate time-windowed sequences -----
        df_used = df_test_raw[used_vars].copy()
        df_scaled = pd.DataFrame(scaler.transform(df_used), columns=used_vars)

        patterns = []
        valid_indices = []
        for i in range(window_size, len(df_scaled)):
            window = []
            for var in used_vars:
                window.extend(df_scaled[var].iloc[i - window_size:i].values)
            patterns.append(window)
            valid_indices.append(i)

        df_patterns = pd.DataFrame(patterns, columns=used_expanded_cols)

        # ----- Match to nearest centroid -----
        df_centroids_sub = df_centroids[used_expanded_cols]
        distances = euclidean_distances(df_patterns.values, df_centroids_sub.values)
        closest_cluster = np.argmin(distances, axis=1)

        # ----- Infer missing values -----
        inferred_values = []
        for row_idx, cluster_idx in enumerate(closest_cluster):
            centroid_row = df_centroids.iloc[cluster_idx]
            inferred_vals = [centroid_row[col] for col in missing_target_cols]
            inferred_values.append(inferred_vals)

        df_inferred = pd.DataFrame(inferred_values, columns=missing_target_cols)
        df_final = df_test_raw.iloc[valid_indices].reset_index(drop=True)
        df_output = pd.concat([df_final, df_inferred], axis=1)
        df_output["Cluster"] = closest_cluster

        st.subheader("✅ Inferred Values Preview")
        st.write(df_output.head())

        # ----- Write to Excel -----
        wb = load_workbook(test_file)
        if "INN_Inference" in wb.sheetnames:
            del wb["INN_Inference"]
        ws = wb.create_sheet("INN_Inference")
        for r in dataframe_to_rows(df_output, index=False, header=True):
            ws.append(r)
        wb.save(output_path)
        st.success(f"📁 Output saved to: {output_path}")

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")

