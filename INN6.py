import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.linear_model import LinearRegression
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import joblib
import io
import base64

# --- Quantum Clustering Functions ---
def quantum_distance(point, centroid, alpha, beta, gamma, weights):
    weighted_diff = weights * np.abs(point - centroid)
    base_dist = np.linalg.norm(weighted_diff)
    exp_term = np.exp(-alpha * base_dist)
    inv_term = beta / (1 + gamma * base_dist)
    return exp_term + inv_term

def update_centroids(clusters, data, gamma, kappa=0.1):
    centroids = []
    for cluster in clusters:
        if cluster:
            mean = np.mean(data[cluster], axis=0)
            interaction = sum([
                np.exp(-kappa * np.linalg.norm(mean - np.mean(data[other], axis=0))) * np.mean(data[other], axis=0)
                for other in range(len(clusters)) if other != clusters.index(cluster)
            ])
            interaction /= (len(clusters) - 1)
            new_centroid = gamma * mean + (1 - gamma) * interaction
            centroids.append(new_centroid)
        else:
            centroids.append(np.zeros(data.shape[1]))
    return np.array(centroids)

def quantum_clustering(data, n_clusters, alpha=1.0, beta=0.5, gamma=0.5, kappa=0.1, tol=1e-4, max_iter=50):
    np.random.seed(42)
    idx = np.random.choice(len(data), n_clusters, replace=False)
    centroids = data[idx]
    weights = np.ones(data.shape[1])
    for _ in range(max_iter):
        clusters = [[] for _ in range(n_clusters)]
        for i, point in enumerate(data):
            dists = [quantum_distance(point, c, alpha, beta, gamma, weights) for c in centroids]
            clusters[np.argmax(dists)].append(i)
        new_centroids = update_centroids(clusters, data, gamma, kappa)
        if np.max(np.abs(new_centroids - centroids)) < tol:
            break
        centroids = new_centroids
    labels = np.zeros(len(data))
    for idx, cluster in enumerate(clusters):
        for i in cluster:
            labels[i] = idx
    return labels.astype(int), centroids

def assign_cluster(sample, centroids):
    # Euclidean distance assignment for test samples
    dists = [np.linalg.norm(sample - c) for c in centroids]
    return np.argmin(dists)

# --- Utility functions ---
def interpret_level(value):
    if value < 0.33:
        return "low"
    elif value < 0.66:
        return "moderate"
    else:
        return "high"

def generate_cluster_descriptions(centroids, feature_names):
    descriptions = []
    for i, centroid in enumerate(centroids):
        traits = []
        for feat_name, val in zip(feature_names, centroid):
            level = interpret_level(val)
            traits.append(f"{level} {feat_name}")
        descriptions.append(f"Cluster {i} is characterized by {', '.join(traits)}.")
    return descriptions

def generate_comparative_summary(centroids, feature_names):
    ranges = centroids.max(axis=0) - centroids.min(axis=0)
    important_idx = np.where(ranges > 0.2)[0]
    if len(important_idx) == 0:
        return "Clusters show similar profiles."
    lines = ["Comparative summary of cluster differences:"]
    for idx in important_idx:
        feat = feature_names[idx]
        vals = centroids[:, idx]
        highs = np.where(vals > 0.66)[0]
        lows = np.where(vals < 0.33)[0]
        lines.append(f"- '{feat}' is high in clusters {list(highs)} and low in clusters {list(lows)}")
    return "\n".join(lines)

def load_scaler_from_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    if "Scaler" not in wb.sheetnames:
        raise ValueError("Scaler sheet not found in Excel file.")
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(row[0]) for row in scaler_sheet.iter_rows(values_only=True) if row[0]]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    scaler = joblib.load(io.BytesIO(scaler_bytes))
    return scaler

def create_lagged_features(df, features, window_size):
    lagged = {}
    for feat in features:
        for lag in range(window_size):
            lagged[f"{feat}_t{lag}"] = df[feat].shift(window_size - 1 - lag)
    lagged_df = pd.DataFrame(lagged).dropna().reset_index(drop=True)
    return lagged_df

# --- Streamlit App ---
st.set_page_config(page_title="CNN-EQIC + INN", layout="wide")
st.title("📊 CNN-EQIC Clustering & 🧠 INN Inference")
tabs = st.tabs(["CNN-EQIC Clustering", "INN Inference"])

with tabs[0]:
    st.header("CNN-EQIC Clustering: Hybrid Neural Network Clustering")
    upload_train = st.file_uploader("Upload training dataset (CSV or Excel)", type=["csv", "xlsx"], key="train")
    if upload_train:
        df = pd.read_csv(upload_train) if upload_train.name.endswith(".csv") else pd.read_excel(upload_train)
        st.dataframe(df.head())

        numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        default_selection = numerical_cols[:min(4, len(numerical_cols))]
        selected = st.multiselect("Select features for clustering", numerical_cols, default=default_selection)

        window_size = st.slider("Window size (lag)", 2, 20, 5)
        if len(selected) >= 2:
            clean = SimpleImputer().fit_transform(df[selected])
            scaler = MinMaxScaler().fit(clean)
            scaled = scaler.transform(clean)

            # Create lagged patterns
            patterns = []
            for i in range(window_size, len(scaled)):
                pattern = scaled[i - window_size:i].flatten()
                patterns.append(pattern)
            patterns = np.array(patterns)

            feature_names = [f"{feat}_t{t}" for t in range(window_size) for feat in selected]

            k = st.slider("Number of clusters", 2, min(10, len(patterns)), 5)
            kmeans = KMeans(n_clusters=k, random_state=42).fit(patterns)
            labels = kmeans.labels_
            centroids = kmeans.cluster_centers_

            st.metric("Silhouette", f"{silhouette_score(patterns, labels):.3f}")
            st.metric("Davies-Bouldin", f"{davies_bouldin_score(patterns, labels):.3f}")
            st.metric("Calinski-Harabasz", f"{calinski_harabasz_score(patterns, labels):.1f}")

            st.subheader("PCA Cluster Projection")
            pca = PCA(n_components=2).fit_transform(patterns)
            df_pca = pd.DataFrame(pca, columns=["PC1", "PC2"])
            df_pca["Cluster"] = labels
            fig, ax = plt.subplots()
            sns.scatterplot(data=df_pca, x="PC1", y="PC2", hue="Cluster", palette="tab10", ax=ax)
            st.pyplot(fig)

            descs = generate_cluster_descriptions(centroids, feature_names)
            st.subheader("Cluster Descriptions")
            for desc in descs:
                st.markdown(f"- {desc}")

            st.subheader("Comparative Summary")
            st.text(generate_comparative_summary(centroids, feature_names))

            if st.button("Export Results to Excel"):
                wb = openpyxl.Workbook()
                ws1 = wb.active
                ws1.title = "Cluster Descriptions"
                for line in descs:
                    ws1.append([line])
                ws2 = wb.create_sheet("Centroids")
                df_centroids = pd.DataFrame(centroids, columns=feature_names)
                for r in dataframe_to_rows(df_centroids, index=False, header=True):
                    ws2.append(r)
                ws3 = wb.create_sheet("Cluster Labels")
                df_labels = pd.DataFrame({"Index": range(len(labels)), "Cluster": labels})
                for r in dataframe_to_rows(df_labels, index=False, header=True):
                    ws3.append(r)

                scaler_io = io.BytesIO()
                joblib.dump(scaler, scaler_io)
                scaler_b64 = base64.b64encode(scaler_io.getvalue()).decode("utf-8")
                ws4 = wb.create_sheet("Scaler")
                chunk_size = 1000
                for i in range(0, len(scaler_b64), chunk_size):
                    ws4.append([scaler_b64[i:i + chunk_size]])

                excel_io = io.BytesIO()
                wb.save(excel_io)
                excel_io.seek(0)
                st.download_button("Download Clustering Results", data=excel_io,
                                   file_name="CNN_EQIC_output.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tabs[1]:
    st.header("INN Inference: Missing Variable Imputation")

    cnn_eqic_file = st.file_uploader("Upload CNN-EQIC Excel Output", type=["xlsx"])
    raw_test_file = st.file_uploader("Upload Raw Test Data", type=["csv", "xlsx"])
    output_path = st.text_input("Full path to save inference output (e.g., C:/.../INNoutput.xlsx)")
    window_size = st.number_input("CNN-EQIC window size", min_value=2, max_value=20, value=5)

    clustering_method = st.selectbox("Choose clustering method", ["KMeans (default)", "Quantum-inspired"])

    if cnn_eqic_file and raw_test_file and output_path:
        try:
            df_centroids = pd.read_excel(cnn_eqic_file, sheet_name="Centroids")
            scaler = load_scaler_from_excel(cnn_eqic_file)

            df_test = pd.read_csv(raw_test_file) if raw_test_file.name.endswith(".csv") else pd.read_excel(raw_test_file)

            base_features = sorted(set(col.rsplit("_t", 1)[0] for col in df_centroids.columns))

            missing_candidates = [feat for feat in base_features if feat not in df_test.columns]
            missing_var = None
            if missing_candidates:
                missing_var = missing_candidates[0]
                st.info(f"Automatically detected missing variable to infer: {missing_var}")
            else:
                missing_var = st.selectbox("Select variable to infer (must be missing in test data):", base_features)

            required_inputs = [feat for feat in base_features if feat != missing_var]
            missing_required = [feat for feat in required_inputs if feat not in df_test.columns]
            if missing_required:
                st.error(f"❌ Missing required variables: {missing_required}")
                st.stop()

            df_test_lagged = create_lagged_features(df_test, base_features, window_size)

            used_cols = [col for col in df_centroids.columns if not col.startswith(missing_var + "_t")]
            target_cols = [col for col in df_centroids.columns if col.startswith(missing_var + "_t")]

            test_scaled = scaler.transform(df_test_lagged[used_cols])
            train_centroids_scaled = df_centroids[used_cols].values

            if clustering_method == "Quantum-inspired":
                n_clusters = st.slider("Number of clusters", 2, min(10, len(train_centroids_scaled)), 5)
                alpha = st.number_input("Alpha parameter", value=1.0)
                beta = st.number_input("Beta parameter", value=0.5)
                gamma = st.number_input("Gamma parameter", value=0.5)

                train_labels, centroids = quantum_clustering(
                    train_centroids_scaled,
                    n_clusters=n_clusters,
                    alpha=alpha,
                    beta=beta,
                    gamma=gamma
                )
                st.write(f"Quantum clustering done with {n_clusters} clusters.")
            else:
                k = st.slider("Number of clusters", 2, min(10, len(train_centroids_scaled)), 5)
                kmeans = KMeans(n_clusters=k, random_state=42).fit(train_centroids_scaled)
                train_labels = kmeans.labels_
                centroids = kmeans.cluster_centers_
                st.write(f"KMeans clustering done with {k} clusters.")

            test_labels = np.array([assign_cluster(sample, centroids) for sample in test_scaled])

            inferred_data = df_test_lagged.copy()

            # Simple heuristic: if missing var in test is binary or clustering method quantum, infer with cluster means
            is_binary = False
            if missing_var in df_test.columns:
                unique_vals = df_test[missing_var].dropna().unique()
                is_binary = len(unique_vals) <= 2

            if is_binary or clustering_method == "Quantum-inspired":
                cluster_means = []
                for cluster_idx in range(len(centroids)):
                    idxs = np.where(train_labels == cluster_idx)[0]
                    cluster_mean_vals = df_centroids.iloc[idxs][target_cols].mean()
                    cluster_means.append(cluster_mean_vals.values)
                cluster_means = np.array(cluster_means)

                inferred_vals = np.array([cluster_means[label] for label in test_labels])

                for i, col in enumerate(target_cols):
                    inferred_data[col] = inferred_vals[:, i]

                st.success("Inference done using cluster means.")
            else:
                for target in target_cols:
                    model = LinearRegression()
                    model.fit(df_centroids[used_cols], df_centroids[target])
                    inferred_data[target] = model.predict(df_test_lagged[used_cols])

                st.success("Inference done using lagged feature regression.")

            inferred_scaled = inferred_data[target_cols].values
            # For unscaling: Append inferred columns to used cols, inverse_transform on full set, then extract inferred cols
            full_scaled_for_inverse = np.hstack([df_test_lagged[used_cols].values, inferred_scaled])
            unscaled_full = scaler.inverse_transform(
                np.hstack([df_test_lagged[used_cols].values, inferred_scaled])
            )
            inferred_unscaled = unscaled_full[:, -len(target_cols):]

            for i, col in enumerate(target_cols):
                inferred_data[col + "_unscaled"] = inferred_unscaled[:, i]

            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Raw Test Data"
            for r in dataframe_to_rows(df_test, index=False, header=True):
                ws1.append(r)

            ws2 = wb.create_sheet("INN_Inference")
            for r in dataframe_to_rows(inferred_data, index=False, header=True):
                ws2.append(r)

            ws3 = wb.create_sheet("Test_Clusters")
            df_clusters = pd.DataFrame({"Index": df_test_lagged.index, "Cluster": test_labels})
            for r in dataframe_to_rows(df_clusters, index=False, header=True):
                ws3.append(r)

            wb.save(output_path)
            st.success(f"📁 Output saved: {output_path}")

        except Exception as e:
            st.error(f"❌ Error: {e}")
