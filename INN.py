# ---------------- INN: Inference Neural Network ----------------
import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import Ridge
from sklearn.metrics.pairwise import euclidean_distances
from openpyxl import load_workbook
import re

st.set_page_config(page_title="INN - Cluster Inference Engine", layout="wide")
st.title("🧠 INN: Inference Neural Network from CNN-EQIC Clusters")

# --- Load CNN-EQIC Excel File ---
st.subheader("Step 1: Load CNN-EQIC Cluster Excel File")
cluster_file = st.file_uploader("Upload CNN-EQIC Excel Output", type="xlsx")

if cluster_file:
    wb = load_workbook(cluster_file, data_only=True)
    sheetnames = wb.sheetnames
    if "Centroids" not in sheetnames:
        st.error("Invalid file: 'Centroids' sheet not found.")
    else:
        df_centroids = pd.read_excel(cluster_file, sheet_name="Centroids")
        cluster_count = df_centroids.shape[0]

        # Detect available variables
        feature_names = df_centroids.columns.tolist()
        base_vars = sorted(set(re.sub(r'_t\d+$', '', f) for f in feature_names))

        # --- Load New Data with Missing Variable ---
        st.subheader("Step 2: Upload New Input Data (with 1 missing variable)")
        new_data_file = st.file_uploader("Upload New Excel or CSV File", type=["xlsx", "csv"], key="newdata")

        if new_data_file:
            df_new = pd.read_csv(new_data_file) if new_data_file.name.endswith(".csv") else pd.read_excel(new_data_file)
            df_new = df_new.copy()

            st.subheader("Step 3: Select Variable to Infer")
            possible_missing = [col for col in base_vars if all(f"{col}_t0" not in df_new.columns for col in [col])]
            selected_missing = st.selectbox("Select the missing variable to infer:", base_vars)

            if selected_missing:
                # Build expanded columns for selected missing variable
                missing_cols = [f for f in feature_names if f.startswith(selected_missing + "_t")]
                input_cols = [col for col in feature_names if col not in missing_cols]

                if all(col in df_new.columns for col in input_cols):
                    st.success(f"✅ All input columns found for inference: {input_cols}")
                    
                    # Step 1: Match new data to closest cluster centroid
                    X_input = df_new[input_cols].values
                    X_centroids = df_centroids[input_cols].values
                    distances = euclidean_distances(X_input, X_centroids)
                    best_clusters = np.argmin(distances, axis=1)

                    # Step 2: Load raw data from matched clusters and train model
                    predicted_values = []
                    for i, row in enumerate(X_input):
                        cluster_idx = best_clusters[i]
                        sheet_name = f"Cluster_{cluster_idx}_RawData"
                        if sheet_name not in sheetnames:
                            st.error(f"Missing sheet: {sheet_name} in Excel file.")
                            continue
                        df_cluster = pd.read_excel(cluster_file, sheet_name=sheet_name)

                        if all(col in df_cluster.columns for col in input_cols + missing_cols):
                            model = Ridge(alpha=1.0)
                            X_train = df_cluster[input_cols].values
                            y_train = df_cluster[missing_cols].mean(axis=1).values
                            model.fit(X_train, y_train)
                            pred = model.predict([row])[0]
                            predicted_values.append(pred)
                        else:
                            st.warning(f"Missing required columns in {sheet_name}.")
                            predicted_values.append(np.nan)

                    df_new[f"{selected_missing}_inferred"] = predicted_values

                    st.subheader("📈 Inference Results")
                    st.dataframe(df_new[[*input_cols, f"{selected_missing}_inferred"]])

                    st.download_button("📥 Download Inference Result as CSV",
                                       data=df_new.to_csv(index=False).encode('utf-8'),
                                       file_name="inference_result.csv",
                                       mime="text/csv")
                else:
                    st.error("Some input columns are missing in your new data.")
