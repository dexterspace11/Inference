# ---------------- Inverse Transformer App ----------------
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import base64
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Inverse Transformer", layout="wide")
st.title("🔁 Scaler-Based Inverse Transformer for CNN-EQIC Outputs")

# ---------------- Helper: Load Scaler from Excel Sheet ----------------
def load_scaler_from_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    if "Scaler" not in wb.sheetnames:
        raise ValueError("Scaler sheet not found in Excel file.")
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(row[0]) for row in scaler_sheet.iter_rows(values_only=True) if row[0]]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    scaler = joblib.load(io.BytesIO(scaler_bytes))
    return scaler

# ---------------- Upload Files ----------------
st.header("Step 1: Upload Excel File with Scaled Data (from CNN-EQIC)")
excel_file = st.file_uploader("Upload Excel file with scaled values and Scaler sheet", type=["xlsx"])

if excel_file:
    sheet_names = pd.ExcelFile(excel_file).sheet_names
    st.write("Detected Sheets:", sheet_names)

    selected_sheet = st.selectbox("Select sheet to inverse transform", sheet_names)

    df_scaled = pd.read_excel(excel_file, sheet_name=selected_sheet)
    st.dataframe(df_scaled.head())

    if st.button("Inverse Transform"):
        try:
            scaler = load_scaler_from_excel(excel_file)
            # Match scaler's expected feature count
            if df_scaled.shape[1] != scaler.n_features_in_:
                st.warning(f"⚠️ Column mismatch: Scaler expects {scaler.n_features_in_} features, "
                           f"but sheet has {df_scaled.shape[1]}. Check column alignment.")
            else:
                unscaled_values = scaler.inverse_transform(df_scaled)
                df_unscaled = pd.DataFrame(unscaled_values, columns=df_scaled.columns)
                st.subheader("✅ Inverse Transformed Data")
                st.write(df_unscaled.head())

                # Save result
                result_path = st.text_input("Enter path to save Excel output", "C:/path/to/UnscaledOutput.xlsx")
                if result_path and st.button("Export to Excel"):
                    with pd.ExcelWriter(result_path, engine="openpyxl") as writer:
                        df_scaled.to_excel(writer, sheet_name="Scaled", index=False)
                        df_unscaled.to_excel(writer, sheet_name="Unscaled", index=False)
                    st.success(f"🎉 File saved to: {result_path}")
        except Exception as e:
            st.error(f"❌ Failed to inverse transform: {e}")
