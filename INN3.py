# ---------------- Inference Neural Network (INN) Engine for Cluster-Based Variable Imputation ----------------
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import base64
import io
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

st.set_page_config(page_title="INN - Cluster-Based Inference Engine", layout="wide")
st.title("🧠 INN - Inference Neural Network for Missing Variable Imputation")

# ---------------- Helper: Load Scaler from 'Scaler' sheet ----------------
def load_scaler_from_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    if "Scaler" not in wb.sheetnames:
        raise ValueError("Scaler sheet not found in Excel file.")
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(cell[0].value) for cell in scaler_sheet.iter_rows(values_only=True) if cell[0].value]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    scaler = joblib.load(io.BytesIO(scaler_bytes))
    return scaler

# ---------------- File Upload ----------------
st.header("Step 1: Upload CNN-EQIC Excel Output")
cluster_file = st.file_uploader("Upload CNN-EQIC Excel output (with full variable data)", type=["xlsx"])

st.header("Step 2: Upload Data with Missing Variable")
missing_file = st.file_uploader("Upload new data with one missing variable", type=["xlsx"])

st.header("Step 3: Enter Output File Path")
output_path = st.text_input("Enter full path to save inference output (e.g., C:/path/to/INNoutput.xlsx)")

if cluster_file and missing_file and output_path:
    try:
        # Load centroids sheet as training data
        df_centroids = pd.read_excel(cluster_file, sheet_name="Centroids")

        # Load new data to infer
        df_input = pd.read_excel(missing_file)

        expected_columns = set(df_centroids.columns)
        input_columns = set(df_input.columns)

        # Identify missing columns
        missing_columns = expected_columns - input_columns
        missing_bases = set([col.split("_t")[0] for col in missing_columns])

        if len(missing_bases) != 1:
            st.error(f"❌ Expected exactly 1 missing variable, but found: {missing_bases}")
            st.stop()

        missing_var = list(missing_bases)[0]
        st.success(f"🔍 INN is inferring missing variable: {missing_var}")

        # Columns used for training and prediction
        used_columns = [col for col in df_centroids.columns if not col.startswith(missing_var + '_')]
        target_columns = [col for col in df_centroids.columns if col.startswith(missing_var + '_')]

        missing_inputs = set(used_columns) - set(df_input.columns)
        if missing_inputs:
            st.error(f"❌ Required input columns are missing from your new data: {missing_inputs}")
            st.stop()

        # Train model for each time step of the missing variable
        inferred_data = df_input.copy()
        for target_col in target_columns:
            model = LinearRegression()
            model.fit(df_centroids[used_columns], df_centroids[target_col])
            inferred_data[target_col] = model.predict(df_input[used_columns])

        st.subheader("✅ Inferred Values (Scaled)")
        st.write(inferred_data[target_columns].head())

        # Load and apply inverse transform using saved scaler
        try:
            scaler = load_scaler_from_excel(cluster_file)
            inferred_scaled = inferred_data[target_columns].values
            inferred_unscaled = scaler.inverse_transform(inferred_scaled)
            for i, col in enumerate(target_columns):
                inferred_data[col + "_unscaled"] = inferred_unscaled[:, i]
            st.subheader("🎯 Inferred Values (Inverse Transformed)")
            st.write(inferred_data[[col + "_unscaled" for col in target_columns]].head())
        except Exception as scaler_err:
            st.warning(f"⚠️ Inverse transform skipped: {scaler_err}")

        # Save to Excel
        wb = load_workbook(missing_file)
        if "INN_Inference" in wb.sheetnames:
            del wb["INN_Inference"]
        ws = wb.create_sheet("INN_Inference")
        for r in dataframe_to_rows(inferred_data, index=False, header=True):
            ws.append(r)
        wb.save(output_path)
        st.success(f"📁 Inference result saved to: {output_path}")

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")
