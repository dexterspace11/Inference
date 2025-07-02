# ---------------- Streamlit INN Calculator Tab ----------------
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import openpyxl
import io
import base64
from sklearn.linear_model import LinearRegression
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- Utility Functions ----------------
def load_scaler_from_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    if "Scaler" not in wb.sheetnames:
        raise ValueError("Scaler sheet not found in Excel file.")
    scaler_sheet = wb["Scaler"]
    b64_chunks = [str(row[0]) for row in scaler_sheet.iter_rows(values_only=True) if row[0]]
    scaler_b64 = "".join(b64_chunks)
    scaler_bytes = base64.b64decode(scaler_b64)
    scaler = joblib.load(io.BytesIO(scaler_bytes))
    return scaler

# ---------------- INN Inference Calculator Tab ----------------
st.set_page_config(page_title="INN Calculator", layout="wide")
st.title("🧠 INN Calculator for Single-Row Inference")

uploaded_excel = st.file_uploader("Upload CNN-EQIC Excel Output", type=["xlsx"])
window_size = st.number_input("Window size used during clustering", min_value=2, max_value=20, value=5)

if uploaded_excel:
    try:
        df_centroids = pd.read_excel(uploaded_excel, sheet_name="Centroids")
        scaler = load_scaler_from_excel(uploaded_excel)

        # Infer all base features from centroid column headers
        all_columns = df_centroids.columns.tolist()
        base_features = sorted(set(col.rsplit("_t", 1)[0] for col in all_columns))

        target_variable = st.selectbox("Select variable to infer:", base_features)
        input_features = [feat for feat in base_features if feat != target_variable]

        st.subheader("Enter Raw Input Values for Required Variables")
        user_input = {}
        for feat in input_features:
            values = st.text_input(f"Enter {window_size} raw values for '{feat}' (comma-separated)", key=feat)
            try:
                val_list = [float(x.strip()) for x in values.split(",") if x.strip()]
                if len(val_list) != window_size:
                    st.warning(f"'{feat}' needs exactly {window_size} values.")
                user_input[feat] = val_list
            except:
                st.warning(f"Invalid input for {feat}. Please enter numbers only.")

        if all(len(v) == window_size for v in user_input.values()):
            # Construct single row DataFrame with lagged features
            lagged_row = {}
            for feat in input_features:
                for t in range(window_size):
                    lagged_row[f"{feat}_t{t}"] = user_input[feat][t]

            df_input = pd.DataFrame([lagged_row])
            df_input_scaled = pd.DataFrame(scaler.transform(df_input), columns=df_input.columns)

            used_cols = [col for col in df_centroids.columns if not col.startswith(target_variable + "_t")]
            target_cols = [col for col in df_centroids.columns if col.startswith(target_variable + "_t")]

            predicted_scaled = {}
            predicted_unscaled = {}

            for target in target_cols:
                model = LinearRegression()
                model.fit(df_centroids[used_cols], df_centroids[target])
                y_pred = model.predict(df_input_scaled[used_cols])
                predicted_scaled[target] = y_pred[0]

            # Reconstruct matrix to apply inverse transform
            pred_df = pd.DataFrame([predicted_scaled])
            combined_scaled = pd.concat([df_input_scaled, pred_df], axis=1)
            inverse_scaled = scaler.inverse_transform(combined_scaled)
            final_output = pd.DataFrame([inverse_scaled[0]], columns=combined_scaled.columns)

            st.subheader("📈 Inference Result")
            st.dataframe(final_output[[col for col in final_output.columns if col.startswith(target_variable)]])

    except Exception as e:
        st.error(f"\u274c Error: {e}")
